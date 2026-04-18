from __future__ import annotations

import json
import sqlite3
from datetime import timedelta
from pathlib import Path
from typing import Any

import bcrypt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ms_rehab_game.settings import ACHIEVEMENT_KEYS, DB_PATH, DEFAULT_USER_GAME_SETTINGS, medal_for_score


class DatabaseManager:
    def __init__(self, db_path: str | Path = DB_PATH) -> None:
        self.db_path = str(db_path)
        Path(self.db_path).parent.mkdir(parents=True, exist_ok=True)
        self.initialize()

    def connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def initialize(self) -> None:
        with self.connect() as conn:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
                CREATE TABLE IF NOT EXISTS game_sessions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER REFERENCES users(id),
                    game_name TEXT NOT NULL,
                    level INTEGER NOT NULL,
                    cognitive_mode TEXT,
                    controller_hand TEXT,
                    score INTEGER NOT NULL,
                    accuracy REAL,
                    duration_seconds INTEGER,
                    correct_actions INTEGER,
                    total_actions INTEGER,
                    played_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
                CREATE TABLE IF NOT EXISTS achievements (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER REFERENCES users(id),
                    achievement_key TEXT NOT NULL,
                    unlocked_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
                CREATE TABLE IF NOT EXISTS paused_sessions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER,
                    game_name TEXT,
                    level INTEGER,
                    cognitive_mode TEXT,
                    score INTEGER,
                    time_remaining INTEGER,
                    state_json TEXT,
                    saved_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
                CREATE TABLE IF NOT EXISTS user_game_settings (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    game_name TEXT NOT NULL,
                    controller_hand TEXT NOT NULL DEFAULT 'right',
                    duration_minutes INTEGER NOT NULL DEFAULT 3,
                    sound_enabled INTEGER NOT NULL DEFAULT 1,
                    cognitive_mode TEXT NOT NULL DEFAULT 'calm',
                    show_tutorial INTEGER NOT NULL DEFAULT 1,
                    UNIQUE(user_id, game_name)
                );
                """
            )

    def create_user(self, username: str, password: str) -> tuple[bool, str]:
        password_hash = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
        try:
            with self.connect() as conn:
                cursor = conn.execute(
                    "INSERT INTO users (username, password_hash) VALUES (?, ?)",
                    (username.strip(), password_hash),
                )
                user_id = cursor.lastrowid
            for game_name in ("thumb_tango", "mindful_tower"):
                self.save_user_game_settings(user_id, game_name, DEFAULT_USER_GAME_SETTINGS.copy())
            return True, "Account created successfully."
        except sqlite3.IntegrityError:
            return False, "That username is already in use."

    def authenticate_user(self, username: str, password: str) -> dict[str, Any] | None:
        with self.connect() as conn:
            row = conn.execute("SELECT * FROM users WHERE username = ?", (username.strip(),)).fetchone()
        if row and bcrypt.checkpw(password.encode("utf-8"), row["password_hash"].encode("utf-8")):
            return dict(row)
        return None

    def get_user_game_settings(self, user_id: int, game_name: str) -> dict[str, Any]:
        with self.connect() as conn:
            row = conn.execute(
                "SELECT * FROM user_game_settings WHERE user_id = ? AND game_name = ?",
                (user_id, game_name),
            ).fetchone()
        if row:
            return {
                "controller_hand": row["controller_hand"],
                "duration_minutes": row["duration_minutes"],
                "sound_enabled": bool(row["sound_enabled"]),
                "cognitive_mode": row["cognitive_mode"],
                "show_tutorial": bool(row["show_tutorial"]),
            }
        self.save_user_game_settings(user_id, game_name, DEFAULT_USER_GAME_SETTINGS.copy())
        return DEFAULT_USER_GAME_SETTINGS.copy()

    def save_user_game_settings(self, user_id: int, game_name: str, settings: dict[str, Any]) -> None:
        merged = DEFAULT_USER_GAME_SETTINGS.copy()
        merged.update(settings)
        with self.connect() as conn:
            conn.execute(
                """
                INSERT INTO user_game_settings (
                    user_id, game_name, controller_hand, duration_minutes, sound_enabled, cognitive_mode, show_tutorial
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(user_id, game_name) DO UPDATE SET
                    controller_hand=excluded.controller_hand,
                    duration_minutes=excluded.duration_minutes,
                    sound_enabled=excluded.sound_enabled,
                    cognitive_mode=excluded.cognitive_mode,
                    show_tutorial=excluded.show_tutorial
                """,
                (
                    user_id,
                    game_name,
                    merged["controller_hand"],
                    int(merged["duration_minutes"]),
                    1 if merged["sound_enabled"] else 0,
                    merged["cognitive_mode"],
                    1 if merged["show_tutorial"] else 0,
                ),
            )

    def save_paused_session(
        self,
        user_id: int,
        game_name: str,
        level: int,
        cognitive_mode: str,
        score: int,
        time_remaining: int,
        state: dict[str, Any],
    ) -> None:
        with self.connect() as conn:
            conn.execute("DELETE FROM paused_sessions WHERE user_id = ? AND game_name = ?", (user_id, game_name))
            conn.execute(
                """
                INSERT INTO paused_sessions (
                    user_id, game_name, level, cognitive_mode, score, time_remaining, state_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (user_id, game_name, level, cognitive_mode, score, time_remaining, json.dumps(state)),
            )

    def get_paused_session(self, user_id: int, game_name: str) -> dict[str, Any] | None:
        with self.connect() as conn:
            row = conn.execute(
                "SELECT * FROM paused_sessions WHERE user_id = ? AND game_name = ? ORDER BY saved_at DESC LIMIT 1",
                (user_id, game_name),
            ).fetchone()
        if not row:
            return None
        payload = dict(row)
        payload["state_json"] = json.loads(payload["state_json"]) if payload["state_json"] else {}
        return payload

    def clear_paused_session(self, user_id: int, game_name: str) -> None:
        with self.connect() as conn:
            conn.execute("DELETE FROM paused_sessions WHERE user_id = ? AND game_name = ?", (user_id, game_name))

    def save_session(
        self,
        user_id: int,
        game_name: str,
        level: int,
        cognitive_mode: str,
        controller_hand: str,
        score: int,
        accuracy: float,
        duration_seconds: int,
        correct_actions: int,
        total_actions: int,
        meta: dict[str, Any] | None = None,
    ) -> tuple[int, list[str]]:
        with self.connect() as conn:
            cursor = conn.execute(
                """
                INSERT INTO game_sessions (
                    user_id, game_name, level, cognitive_mode, controller_hand, score,
                    accuracy, duration_seconds, correct_actions, total_actions
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    user_id,
                    game_name,
                    level,
                    cognitive_mode,
                    controller_hand,
                    score,
                    accuracy,
                    duration_seconds,
                    correct_actions,
                    total_actions,
                ),
            )
            session_id = cursor.lastrowid
        achievements = self.evaluate_achievements(
            user_id,
            game_name,
            score,
            perfect=total_actions > 0 and correct_actions == total_actions,
            streak=max((meta or {}).get("best_streak", 0), 0),
        )
        self.clear_paused_session(user_id, game_name)
        return session_id, achievements

    def get_sessions_dataframe(self, user_id: int, game_name: str | None = None) -> pd.DataFrame:
        query = "SELECT * FROM game_sessions WHERE user_id = ?"
        params: list[Any] = [user_id]
        if game_name:
            query += " AND game_name = ?"
            params.append(game_name)
        query += " ORDER BY played_at ASC"
        with self.connect() as conn:
            return pd.read_sql_query(query, conn, params=params, parse_dates=["played_at"])

    def get_statistics_summary(self, user_id: int, game_name: str | None = None) -> dict[str, Any]:
        df = self.get_sessions_dataframe(user_id, game_name)
        if df.empty:
            return {
                "games_played": 0,
                "avg_accuracy": 0.0,
                "best_score": 0,
                "days_active": 0,
                "total_therapy_minutes": 0,
                "avg_session_minutes": 0.0,
                "current_streak_days": 0,
                "trend_accuracy": "N/A",
                "trend_label": "Insufficient data",
                "trend_color": "neutral",
                "top_hand": "N/A",
                "top_level": "N/A",
                "compliance_rate": 0.0,
            }

        total_sessions = len(df)
        avg_accuracy = float(df["accuracy"].fillna(0).mean())
        best_score = int(df["score"].max())
        days_active = int(df["played_at"].dt.date.nunique())
        total_minutes = int(df["duration_seconds"].fillna(0).sum() // 60)
        avg_session_min = round(df["duration_seconds"].fillna(0).mean() / 60, 1)

        # Adherence / compliance: sessions per active week
        if days_active >= 1:
            df["date"] = df["played_at"].dt.date
            date_range_days = (df["date"].max() - df["date"].min()).days + 1
            weeks = max(date_range_days / 7, 1)
            sessions_per_week = round(total_sessions / weeks, 1)
            # Typical rehab target: 3+ sessions/week
            compliance_rate = min(round((sessions_per_week / 3) * 100, 1), 100.0)
        else:
            sessions_per_week = 0
            compliance_rate = 0.0

        # Accuracy trend: compare last 3 sessions vs previous 3
        acc_series = df["accuracy"].fillna(0).tolist()
        if len(acc_series) >= 6:
            recent = sum(acc_series[-3:]) / 3
            previous = sum(acc_series[-6:-3]) / 3
            delta = recent - previous
            if delta >= 5:
                trend_label = f"Improving (+{delta:.1f}%)"
                trend_color = "good"
            elif delta <= -5:
                trend_label = f"Declining ({delta:.1f}%)"
                trend_color = "warn"
            else:
                trend_label = "Stable"
                trend_color = "neutral"
            trend_accuracy = f"{recent:.1f}%"
        else:
            trend_label = "More data needed"
            trend_color = "neutral"
            trend_accuracy = f"{avg_accuracy:.1f}%"

        # Most used hand and highest reached level
        top_hand = df["controller_hand"].mode()[0].capitalize() if not df["controller_hand"].isna().all() else "N/A"
        top_level = str(int(df["level"].max())) if not df["level"].isna().all() else "N/A"

        # Current consecutive-days streak
        current_streak = self._current_streak(df)

        return {
            "games_played": total_sessions,
            "avg_accuracy": avg_accuracy,
            "best_score": best_score,
            "days_active": days_active,
            "total_therapy_minutes": total_minutes,
            "avg_session_minutes": avg_session_min,
            "sessions_per_week": sessions_per_week,
            "compliance_rate": compliance_rate,
            "current_streak_days": current_streak,
            "trend_accuracy": trend_accuracy,
            "trend_label": trend_label,
            "trend_color": trend_color,
            "top_hand": top_hand,
            "top_level": top_level,
        }

    def _current_streak(self, df: pd.DataFrame) -> int:
        """Return the current (ongoing) consecutive-day play streak."""
        if df.empty:
            return 0
        from datetime import date as date_type
        dates = sorted(set(df["played_at"].dt.date), reverse=True)
        today = date_type.today()
        if dates[0] not in (today, today - timedelta(days=1)):
            return 0
        streak = 1
        for prev, current in zip(dates, dates[1:]):
            if prev - current == timedelta(days=1):
                streak += 1
            else:
                break
        return streak

    def get_clinical_trend_data(self, user_id: int, game_name: str | None = None) -> dict[str, Any]:
        """
        Returns week-by-week aggregates useful for clinical progress monitoring.
        Each entry covers: week label, avg accuracy, total sessions, total minutes,
        avg score, and motor throughput (correct actions per minute).
        """
        df = self.get_sessions_dataframe(user_id, game_name)
        if df.empty:
            return {"weeks": [], "columns": []}

        df["week"] = df["played_at"].dt.to_period("W").apply(lambda p: str(p.start_time.date()))
        weekly = (
            df.groupby("week")
            .agg(
                sessions=("id", "count"),
                avg_accuracy=("accuracy", "mean"),
                total_minutes=("duration_seconds", lambda x: round(x.sum() / 60, 1)),
                avg_score=("score", "mean"),
                correct_actions=("correct_actions", "sum"),
                duration_seconds=("duration_seconds", "sum"),
            )
            .reset_index()
        )
        weekly["motor_throughput"] = (
            weekly["correct_actions"] / (weekly["duration_seconds"] / 60).clip(lower=1)
        ).round(1)
        weekly["avg_accuracy"] = weekly["avg_accuracy"].round(1)
        weekly["avg_score"] = weekly["avg_score"].round(0).astype(int)
        return weekly.to_dict(orient="records")

    def export_sessions_to_excel(self, user_id: int, export_path: str | Path) -> Path:
        """
        Export a doctor-friendly clinical report Excel file with:
        - Summary sheet with KPIs and clinical interpretation
        - Weekly progress sheet
        - Raw sessions sheet
        """
        export_path = Path(export_path)
        export_path.parent.mkdir(parents=True, exist_ok=True)
        df = self.get_sessions_dataframe(user_id)

        wb = Workbook()

        # ── Styles ──────────────────────────────────────────────────────────
        header_fill = PatternFill("solid", fgColor="1A3B5D")
        subheader_fill = PatternFill("solid", fgColor="2C5F8A")
        good_fill = PatternFill("solid", fgColor="D4EDDA")
        warn_fill = PatternFill("solid", fgColor="FFF3CD")
        bad_fill = PatternFill("solid", fgColor="F8D7DA")
        neutral_fill = PatternFill("solid", fgColor="E2E8F0")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, color="1A3B5D", size=14)
        label_font = Font(bold=True, color="1A3B5D", size=10)
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        def style_header_row(ws, row_num, num_cols):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

        def style_data_row(ws, row_num, num_cols, fill=None):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row_num, column=col)
                if fill:
                    cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

        # ── Sheet 1: Clinical Summary ────────────────────────────────────────
        ws_summary = wb.active
        ws_summary.title = "Clinical Summary"
        ws_summary.column_dimensions["A"].width = 32
        ws_summary.column_dimensions["B"].width = 22
        ws_summary.column_dimensions["C"].width = 40

        # Get username from DB
        with self.connect() as conn:
            user_row = conn.execute("SELECT username, created_at FROM users WHERE id = ?", (user_id,)).fetchone()
        username = user_row["username"] if user_row else f"Patient #{user_id}"
        enrolled_date = user_row["created_at"][:10] if user_row else "Unknown"

        # Title block
        ws_summary.merge_cells("A1:C1")
        ws_summary["A1"] = "MS Rehabilitation Game — Clinical Progress Report"
        ws_summary["A1"].font = Font(bold=True, color="FFFFFF", size=15)
        ws_summary["A1"].fill = header_fill
        ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_summary.row_dimensions[1].height = 32

        ws_summary.merge_cells("A2:C2")
        ws_summary["A2"] = f"Patient: {username}   |   Enrolled: {enrolled_date}   |   Report generated: {pd.Timestamp.now().strftime('%Y-%m-%d')}"
        ws_summary["A2"].font = Font(italic=True, color="555555", size=10)
        ws_summary["A2"].alignment = Alignment(horizontal="center")

        ws_summary.append([])  # blank row

        # Section headers
        def write_section(ws, start_row, title, rows_data):
            ws.merge_cells(f"A{start_row}:C{start_row}")
            ws.cell(row=start_row, column=1).value = title
            ws.cell(row=start_row, column=1).font = Font(bold=True, color="FFFFFF", size=11)
            ws.cell(row=start_row, column=1).fill = subheader_fill
            ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[start_row].height = 22
            r = start_row + 1
            for label, value, note, fill in rows_data:
                ws.cell(row=r, column=1).value = label
                ws.cell(row=r, column=1).font = label_font
                ws.cell(row=r, column=1).border = thin_border
                ws.cell(row=r, column=2).value = value
                ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=r, column=2).border = thin_border
                if fill:
                    ws.cell(row=r, column=2).fill = fill
                ws.cell(row=r, column=3).value = note
                ws.cell(row=r, column=3).font = Font(italic=True, color="555555", size=9)
                ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="center")
                ws.cell(row=r, column=3).border = thin_border
                ws.row_dimensions[r].height = 20
                r += 1
            return r + 1  # next section starts after blank row

        if df.empty:
            ws_summary.append(["No session data recorded yet.", "", ""])
        else:
            summary = self.get_statistics_summary(user_id)
            acc = summary["avg_accuracy"]
            compliance = summary["compliance_rate"]
            trend_color = summary["trend_color"]

            acc_fill = good_fill if acc >= 75 else (warn_fill if acc >= 50 else bad_fill)
            compliance_fill = good_fill if compliance >= 80 else (warn_fill if compliance >= 50 else bad_fill)
            trend_fill = good_fill if trend_color == "good" else (warn_fill if trend_color == "warn" else neutral_fill)

            next_row = write_section(ws_summary, 4, "📊 Engagement & Adherence", [
                ("Total Sessions Completed", summary["games_played"], "Number of therapy sessions the patient has completed.", None),
                ("Active Therapy Days", summary["days_active"], "Number of distinct days the patient engaged with the game.", None),
                ("Total Therapy Time (min)", summary["total_therapy_minutes"], "Cumulative active exercise time across all sessions.", None),
                ("Avg. Session Duration (min)", summary["avg_session_minutes"], "Average length of each therapy session.", None),
                ("Sessions per Week (avg)", summary.get("sessions_per_week", "N/A"), "Recommended target: ≥ 3 sessions per week.", None),
                ("Adherence Rate (%)", f"{compliance}%", "How consistently the patient meets the 3 sessions/week target. ≥80% = Good, 50–79% = Fair, <50% = Poor.", compliance_fill),
                ("Current Play Streak (days)", summary["current_streak_days"], "Number of consecutive days with at least one session.", None),
            ])

            next_row = write_section(ws_summary, next_row, "🎯 Motor Performance", [
                ("Average Accuracy (%)", f"{acc:.1f}%", "Proportion of correct gestures vs total attempts. Higher is better.", acc_fill),
                ("Best Score Achieved", summary["best_score"], "Peak session score — reflects highest motor output.", None),
                ("Difficulty Level Reached", summary["top_level"], "Highest difficulty level played (1=Easy, 2=Medium, 3=Hard).", None),
                ("Primary Hand Used", summary["top_hand"], "Hand most frequently used for therapy exercises.", None),
            ])

            next_row = write_section(ws_summary, next_row, "📈 Progress Trend (last 6 sessions)", [
                ("Recent Accuracy (last 3 sessions)", summary["trend_accuracy"], "Average accuracy over the patient's most recent 3 sessions.", trend_fill),
                ("Trend Assessment", summary["trend_label"], "Improving = +5% or more vs previous 3 sessions. Declining = -5% or more.", trend_fill),
            ])

            # Clinical notes
            ws_summary.append([])
            note_row = ws_summary.max_row + 1
            ws_summary.merge_cells(f"A{note_row}:C{note_row}")
            ws_summary.cell(row=note_row, column=1).value = "📝 Clinical Notes"
            ws_summary.cell(row=note_row, column=1).font = Font(bold=True, color="FFFFFF", size=11)
            ws_summary.cell(row=note_row, column=1).fill = subheader_fill
            ws_summary.row_dimensions[note_row].height = 22

            notes = []
            if acc < 50:
                notes.append("• Accuracy below 50% — consider reducing difficulty level or reviewing exercise technique.")
            elif acc >= 85:
                notes.append("• Accuracy above 85% — patient may be ready to progress to a higher difficulty level.")
            if compliance < 50:
                notes.append("• Low adherence rate — consider discussing barriers to regular engagement with the patient.")
            if summary["current_streak_days"] >= 7:
                notes.append("• Patient has maintained a 7+ day streak — positive engagement indicator.")
            if not notes:
                notes.append("• Patient performance is within expected range. Continue current protocol.")

            for note in notes:
                nr = ws_summary.max_row + 1
                ws_summary.merge_cells(f"A{nr}:C{nr}")
                ws_summary.cell(row=nr, column=1).value = note
                ws_summary.cell(row=nr, column=1).font = Font(italic=True, size=10, color="1A3B5D")
                ws_summary.cell(row=nr, column=1).alignment = Alignment(wrap_text=True)
                ws_summary.row_dimensions[nr].height = 20

        # ── Sheet 2: Weekly Progress ─────────────────────────────────────────
        ws_weekly = wb.create_sheet("Weekly Progress")
        weekly_cols = ["Week Starting", "Sessions", "Avg Accuracy (%)", "Total Therapy (min)", "Avg Score", "Motor Throughput (actions/min)"]
        ws_weekly.append(weekly_cols)
        style_header_row(ws_weekly, 1, len(weekly_cols))
        ws_weekly.row_dimensions[1].height = 28

        for col_letter, width in zip(["A","B","C","D","E","F"], [16, 12, 18, 20, 14, 28]):
            ws_weekly.column_dimensions[col_letter].width = width

        weekly_data = self.get_clinical_trend_data(user_id)
        for i, week in enumerate(weekly_data, start=2):
            row = [
                week["week"],
                week["sessions"],
                f"{week['avg_accuracy']}%",
                week["total_minutes"],
                week["avg_score"],
                week["motor_throughput"],
            ]
            ws_weekly.append(row)
            fill = good_fill if i % 2 == 0 else neutral_fill
            style_data_row(ws_weekly, i, len(weekly_cols), fill)

        # ── Sheet 3: Raw Session Log ─────────────────────────────────────────
        ws_raw = wb.create_sheet("Session Log")
        if df.empty:
            ws_raw.append(["No session data available"])
        else:
            # Rename columns for clinical readability
            df_export = df.copy()
            df_export = df_export.rename(columns={
                "id": "Session ID",
                "user_id": "Patient ID",
                "game_name": "Exercise",
                "level": "Difficulty Level",
                "cognitive_mode": "Cognitive Mode",
                "controller_hand": "Hand Used",
                "score": "Score",
                "accuracy": "Accuracy (%)",
                "duration_seconds": "Duration (sec)",
                "correct_actions": "Correct Actions",
                "total_actions": "Total Actions",
                "played_at": "Session Date & Time",
            })
            df_export["Exercise"] = df_export["Exercise"].str.replace("_", " ").str.title()
            df_export["Hand Used"] = df_export["Hand Used"].str.capitalize() if df_export["Hand Used"].dtype == object else df_export["Hand Used"]
            df_export["Accuracy (%)"] = df_export["Accuracy (%)"].round(1)
            df_export["Difficulty Level"] = df_export["Difficulty Level"].apply(lambda x: f"Level {x}")

            ws_raw.append(list(df_export.columns))
            style_header_row(ws_raw, 1, len(df_export.columns))
            ws_raw.row_dimensions[1].height = 28

            for col_idx, col_name in enumerate(df_export.columns, start=1):
                ws_raw.column_dimensions[chr(64 + col_idx)].width = max(len(col_name) + 4, 14)

            for i, row_data in enumerate(df_export.itertuples(index=False), start=2):
                ws_raw.append(list(row_data))
                fill = neutral_fill if i % 2 == 0 else None
                style_data_row(ws_raw, i, len(df_export.columns), fill)

        wb.save(export_path)
        return export_path

    def get_best_score(self, user_id: int, game_name: str) -> int:
        with self.connect() as conn:
            row = conn.execute(
                "SELECT COALESCE(MAX(score), 0) AS best_score FROM game_sessions WHERE user_id = ? AND game_name = ?",
                (user_id, game_name),
            ).fetchone()
        return int(row["best_score"]) if row else 0

    def get_achievements(self, user_id: int) -> set[str]:
        with self.connect() as conn:
            rows = conn.execute("SELECT achievement_key FROM achievements WHERE user_id = ?", (user_id,)).fetchall()
        return {row["achievement_key"] for row in rows}

    def unlock_achievement(self, user_id: int, key: str) -> bool:
        with self.connect() as conn:
            existing = conn.execute(
                "SELECT 1 FROM achievements WHERE user_id = ? AND achievement_key = ?",
                (user_id, key),
            ).fetchone()
            if existing:
                return False
            conn.execute(
                "INSERT INTO achievements (user_id, achievement_key) VALUES (?, ?)",
                (user_id, key),
            )
        return True

    def consecutive_days_played(self, user_id: int) -> int:
        df = self.get_sessions_dataframe(user_id)
        if df.empty:
            return 0
        dates = sorted(set(df["played_at"].dt.date))
        streak = 1
        best = 1
        for prev, current in zip(dates, dates[1:]):
            if current == prev + timedelta(days=1):
                streak += 1
                best = max(best, streak)
            elif current != prev:
                streak = 1
        return best

    def evaluate_achievements(self, user_id: int, game_name: str, score: int, perfect: bool, streak: int) -> list[str]:
        unlocked: list[str] = []
        total_sessions = self.get_sessions_dataframe(user_id)
        if len(total_sessions) == 1 and self.unlock_achievement(user_id, "first_game"):
            unlocked.append("first_game")
        medal = medal_for_score(game_name, score).lower()
        if self.unlock_achievement(user_id, medal):
            unlocked.append(medal)
        for threshold in (5, 10, 15):
            key = f"streak_{threshold}"
            if streak >= threshold and self.unlock_achievement(user_id, key):
                unlocked.append(key)
        if perfect and self.unlock_achievement(user_id, "perfect_game"):
            unlocked.append("perfect_game")
        days = self.consecutive_days_played(user_id)
        for threshold in (10, 20, 30):
            key = f"days_{threshold}"
            if days >= threshold and self.unlock_achievement(user_id, key):
                unlocked.append(key)
        return unlocked

    def available_achievements(self) -> list[str]:
        return ACHIEVEMENT_KEYS